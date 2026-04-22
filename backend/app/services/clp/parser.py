"""
CLP — Input XLSX/PDF Parser Service (v3 — Bulletproof)

Three-layer extraction strategy:
  Layer 1: Anchor-based scanning (handles standard cell layouts)
  Layer 2: Pattern-based scanning (handles merged/unusual layouts)
  Layer 3: LLM-assisted extraction (handles garbled PDFs + scanned images)

Ported from CLP standalone project with updated imports for MySmartStudy.
"""

import re
import asyncio
import datetime as dt_module
from io import BytesIO
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
import pdfplumber

from app.schemas import CLPWeekData as WeekData, CLPUploadMetadata as UploadMetadata
from app.services.clp.gemini_service import extract_file_content, is_exception_week


# ═══════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════

BULAN_MELAYU = {
    1: "Januari", 2: "Februari", 3: "Mac", 4: "April",
    5: "Mei", 6: "Jun", 7: "Julai", 8: "Ogos",
    9: "September", 10: "Oktober", 11: "November", 12: "Disember",
}
BULAN_NAMA_TO_NUM = {name.lower(): num for num, name in BULAN_MELAYU.items()}

_METADATA_ANCHORS: list[tuple[list[str], str]] = [
    (["nama kursus", "mata pelajaran", "course name"], "nama_kursus"),
    (["kod kursus", "course code"], "kod_kursus"),
    (["nama pensyarah", "pensyarah", "lecturer"], "pensyarah"),
    (["kumpulan diajar", "kumpulan ajar"], "kumpulan_diajar"),
    (["program"], "program"),
    (["semester"], "semester"),
    (["tahun"], "tahun"),
    (["ambilan", "intake"], "ambilan"),
    (["jabatan", "unit"], "jabatan"),
    (["jumlah kredit", "kredit"], "jumlah_kredit"),
]

_KOD_KURSUS_PATTERN = re.compile(r'\b([A-Z]{2,5}\d{3,5}[A-Z]?)\b')
_NAMA_PATTERN = re.compile(
    r'\b([A-Z][A-Z .]+\s+(?:BIN(?:TI)?|B\.?|BT\.?)\s+[A-Z][A-Z .]+)\b'
)


# ═══════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def _clean(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = re.sub(r'[\r\n]+', ' ', s)
    s = s.strip(" \t:-;")
    return s


def _cell_text(cell) -> str:
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _matches_anchor(text: str, anchors: list[str]) -> bool:
    t = text.strip().lower().rstrip(":")
    for anchor in anchors:
        if anchor in t:
            return True
    return False


def _is_label_cell(text: str) -> bool:
    t = text.strip().lower()
    for anchors, _ in _METADATA_ANCHORS:
        for a in anchors:
            if a in t:
                return True
    return False


def _split_kumpulan(value: str) -> list[str]:
    groups = re.split(r'[,;&/]', value)
    return [g.strip() for g in groups if g.strip()]


# ═══════════════════════════════════════════════════════════════════════════
#  DATE FORMATTING
# ═══════════════════════════════════════════════════════════════════════════

def _format_tarikh(tarikh: str) -> str:
    if not tarikh:
        return ""

    text = str(tarikh).strip()
    if not text:
        return ""

    text = text.replace("\u2013", "-").replace("\u2014", "-")

    # Pattern 1: "25-29 Ogos 2025"
    m = re.search(r'(\d{1,2})\s*-\s*\d{1,2}\s+([A-Za-z\u00C0-\u00F6\u00F8-\u00FF]+)\s+(\d{4})', text)
    if m:
        day, month_raw, year = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        month_name = BULAN_MELAYU.get(BULAN_NAMA_TO_NUM.get(month_raw.lower(), 0), month_raw)
        return f"{day} {month_name} {year}"

    # Pattern 2: "25 Ogos - 29 Ogos 2025"
    m = re.search(r'(\d{1,2})\s+([A-Za-z\u00C0-\u00F6\u00F8-\u00FF]+)\s*-\s*\d{1,2}\s+[A-Za-z\u00C0-\u00F6\u00F8-\u00FF]+\s+(\d{4})', text)
    if m:
        day, month_raw, year = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        month_name = BULAN_MELAYU.get(BULAN_NAMA_TO_NUM.get(month_raw.lower(), 0), month_raw)
        return f"{day} {month_name} {year}"

    # Pattern 3: "25 Ogos 2025 - 29 Ogos 2025"
    m = re.search(
        r'(\d{1,2})\s+([A-Za-z\u00C0-\u00F6\u00F8-\u00FF]+)\s+(\d{4})\s*-\s*\d{1,2}\s+[A-Za-z\u00C0-\u00F6\u00F8-\u00FF]+\s+\d{4}',
        text,
    )
    if m:
        day, month_raw, year = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        month_name = BULAN_MELAYU.get(BULAN_NAMA_TO_NUM.get(month_raw.lower(), 0), month_raw)
        return f"{day} {month_name} {year}"

    # Pattern 4: "25 Ogos 2025" standalone
    m = re.search(r'(\d{1,2})\s+([A-Za-z\u00C0-\u00F6\u00F8-\u00FF]+)\s+(\d{4})', text)
    if m:
        day, month_raw, year = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        month_num = BULAN_NAMA_TO_NUM.get(month_raw.lower(), 0)
        if month_num:
            return f"{day} {BULAN_MELAYU[month_num]} {year}"

    # Pattern 5: Numeric "25/08/2025"
    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', text)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if year < 100:
            year += 2000
        if 1 <= month <= 12:
            return f"{day} {BULAN_MELAYU[month]} {year}"

    # Pattern 6: ISO format
    try:
        d = dt_module.datetime.fromisoformat(text.replace("Z", ""))
        return f"{d.day} {BULAN_MELAYU[d.month]} {d.year}"
    except (ValueError, TypeError):
        pass

    return text


# ═══════════════════════════════════════════════════════════════════════════
#  LAYER 1: ANCHOR-BASED XLSX EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════

def _extract_metadata_anchored(ws, max_row: int = 20) -> dict[str, str | list[str]]:
    result: dict[str, str | list[str]] = {}

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), values_only=False):
        for cell in row:
            ct = _cell_text(cell)
            if not ct:
                continue

            anchor_count = 0
            for anchors, _ in _METADATA_ANCHORS:
                if _matches_anchor(ct, anchors):
                    anchor_count += 1
            if anchor_count >= 2:
                continue

            for anchors, field_name in _METADATA_ANCHORS:
                if field_name in result:
                    continue
                if not _matches_anchor(ct, anchors):
                    continue

                value = ""

                if ":" in ct:
                    _, _, after = ct.partition(":")
                    after = after.strip()
                    if after and not _is_label_cell(after):
                        value = _clean(after)

                if not value:
                    for c2 in row:
                        if c2.column <= cell.column:
                            continue
                        c2t = _cell_text(c2)
                        if not c2t:
                            continue
                        if _is_label_cell(c2t):
                            continue
                        value = _clean(c2t)
                        break

                if not value:
                    try:
                        below = ws.cell(row=cell.row + 1, column=cell.column)
                        bt = _cell_text(below)
                        if bt and not _is_label_cell(bt):
                            value = _clean(bt)
                    except Exception:
                        pass

                if not value:
                    continue

                if field_name == "kumpulan_diajar":
                    result[field_name] = _split_kumpulan(value)
                else:
                    result[field_name] = value
                break

    return result


# ═══════════════════════════════════════════════════════════════════════════
#  LAYER 2: LABEL-COUNT OFFSET MAPPING FOR MERGED-LABEL LAYOUTS
# ═══════════════════════════════════════════════════════════════════════════

def _extract_metadata_patterns(ws, max_row: int = 20) -> dict[str, str | list[str]]:
    result: dict[str, str | list[str]] = {}

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), values_only=False):
        for cell in row:
            raw_text = cell.value
            if raw_text is None:
                continue
            text = str(raw_text)

            if "\n" not in text and text.count(":") < 2:
                continue

            text_lower = text.lower()
            found_labels: list[tuple[int, str]] = []

            for anchors, field_name in _METADATA_ANCHORS:
                best_pos = -1
                for anchor in anchors:
                    pos = text_lower.find(anchor)
                    if pos >= 0 and (best_pos < 0 or pos < best_pos):
                        best_pos = pos
                if best_pos >= 0:
                    found_labels.append((best_pos, field_name))

            if len(found_labels) < 3:
                continue

            found_labels.sort(key=lambda x: x[0])

            label_row = cell.row
            label_col = cell.column

            value_col = None
            for search_col in range(label_col + 1, min(ws.max_column + 1, label_col + 6)):
                for test_offset in range(min(3, len(found_labels))):
                    try:
                        test_cell = ws.cell(row=label_row + test_offset, column=search_col)
                        if test_cell.value is not None:
                            val = _clean(test_cell.value)
                            if val and not _is_label_cell(val):
                                value_col = search_col
                                break
                    except Exception:
                        pass
                if value_col:
                    break

            if not value_col:
                continue

            for seq_idx, (_, field_name) in enumerate(found_labels):
                if field_name in result:
                    continue
                value_row = label_row + seq_idx
                try:
                    val_cell = ws.cell(row=value_row, column=value_col)
                    if val_cell.value is not None:
                        value = _clean(val_cell.value)
                        if value:
                            if field_name == "kumpulan_diajar":
                                result[field_name] = _split_kumpulan(value)
                            else:
                                result[field_name] = value
                except Exception:
                    pass

    # Fallback regex
    if "kod_kursus" not in result:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                m = _KOD_KURSUS_PATTERN.search(str(cell.value))
                if m:
                    code = m.group(1)
                    full_text = str(cell.value).strip()
                    if full_text == code or len(full_text) < 15:
                        result["kod_kursus"] = code
                        break
            if "kod_kursus" in result:
                break

    if "pensyarah" not in result:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                m = _NAMA_PATTERN.search(text)
                if m and text == m.group(0):
                    result["pensyarah"] = _clean(m.group(1))
                    break
            if "pensyarah" in result:
                break

    return result


# ═══════════════════════════════════════════════════════════════════════════
#  XLSX WEEKLY DATA EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════

def _find_header_row(ws, max_search: int = 25) -> Optional[int]:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_search), values_only=False):
        for cell in row:
            if cell.value is not None:
                val = str(cell.value).strip().lower()
                if val in ("minggu", "week") or "minggu" in val or "inggu" in val:
                    return cell.row
    return None


def _map_columns(ws, header_row: int) -> dict[str, int]:
    col_map: dict[str, int] = {}
    header_anchors = {
        "minggu": ["minggu", "week"],
        "tarikh": ["tarikh", "date"],
        "topik": ["topik", "tajuk", "topic"],
        "jam": ["jam", "hour", "interaksi", "bersemuka", "f2f"],
        "catatan": ["catatan", "refleksi", "nota"],
    }
    for scan_row in range(max(1, header_row - 2), min(ws.max_row + 1, header_row + 3)):
        for cell in ws[scan_row]:
            if cell.value is None:
                continue
            header_text = str(cell.value).strip().lower()
            for field_key, anchor_list in header_anchors.items():
                if field_key in col_map:
                    continue
                if any(a in header_text for a in anchor_list):
                    col_map[field_key] = cell.column
                    break
    return col_map


def _discover_jam_columns(ws, header_row: int) -> dict[str, int]:
    jam_col_map: dict[str, int] = {}
    for sub_row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        row_cells = list(ws.iter_rows(min_row=sub_row, max_row=sub_row, values_only=False))[0]
        ktal_cells = {}
        for c in row_cells:
            if c.value is not None:
                v = str(c.value).strip()
                if v.upper() in ("K", "T", "A", "L") and len(v) == 1:
                    ktal_cells[c.column] = v.upper()
        if len(ktal_cells) >= 4:
            sorted_cols = sorted(ktal_cells.keys())
            for col in sorted_cols[:4]:
                jam_col_map[f"{ktal_cells[col]}(F2F)"] = col
            for col in sorted_cols[4:8]:
                jam_col_map[f"{ktal_cells[col]}(ODL)"] = col
            max_jam = max(sorted_cols) if sorted_cols else 0
            for scan_r in range(header_row, sub_row + 1):
                for cell in ws[scan_r]:
                    if cell.value is None or cell.column <= max_jam:
                        continue
                    v = str(cell.value).strip().lower()
                    if any(kw in v for kw in ["tidak bersemuka", "nf2f", "non face", "(c)"]):
                        jam_col_map["NF2F"] = cell.column
                        break
            break
    return jam_col_map


def _extract_weeks_from_xlsx(ws, header_row: int, col_map: dict[str, int],
                              jam_col_map: dict[str, int]) -> list[WeekData]:
    weeks: list[WeekData] = []
    minggu_col = col_map.get("minggu", 1)

    data_start = None
    for r in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
        val = ws.cell(row=r, column=minggu_col).value
        if val is not None:
            try:
                int(val)
                data_start = r
                break
            except (ValueError, TypeError):
                continue

    if not data_start:
        return weeks

    for row_idx in range(data_start, ws.max_row + 1):
        minggu_val = ws.cell(row=row_idx, column=minggu_col).value
        if minggu_val is None:
            continue
        try:
            week_num = int(minggu_val)
        except (ValueError, TypeError):
            continue
        if week_num < 1 or week_num > 19:
            continue

        tarikh = ""
        if "tarikh" in col_map:
            tarikh_raw = ws.cell(row=row_idx, column=col_map["tarikh"]).value
            if tarikh_raw is not None:
                tarikh = _clean(tarikh_raw)

        topik = ""
        if "topik" in col_map:
            topik_col = col_map["topik"]
            topik_raw = ws.cell(row=row_idx, column=topik_col).value
            if topik_raw is not None:
                topik = _clean(topik_raw)
            if not topik:
                for scan_up in range(1, 4):
                    check_row = row_idx - scan_up
                    if check_row < data_start:
                        break
                    val = ws.cell(row=check_row, column=topik_col).value
                    if val is not None:
                        candidate = _clean(val)
                        if candidate and len(candidate) > 5 and not candidate.isdigit():
                            topik = candidate
                            break

        catatan = ""
        if "catatan" in col_map:
            catatan_raw = ws.cell(row=row_idx, column=col_map["catatan"]).value
            if catatan_raw is not None:
                catatan = _clean(catatan_raw)

        topik_lower = topik.lower()
        if any(kw in topik_lower for kw in ["cuti", "peperiksaan", "ulangkaji", "pertengahan semester"]):
            if "akhir semester" in topik_lower:
                topik = "CUTI AKHIR SEMESTER IPG"
            elif "pertengahan" in topik_lower or "cuti" in topik_lower:
                topik = "CUTI PERTENGAHAN SEMESTER IPG"
            elif "ulangkaji" in topik_lower:
                topik = "MINGGU ULANGKAJI"
            elif "peperiksaan" in topik_lower:
                topik = "PEPERIKSAAN AKHIR"

        jam_parts = []
        for label, col_idx in jam_col_map.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                try:
                    num = int(val) if str(val).strip().replace("`", "").isdigit() else 0
                    if num > 0:
                        jam_parts.append(f"{label}:{num}")
                except (ValueError, TypeError):
                    pass

        weeks.append(WeekData(
            minggu=week_num,
            tarikh=_format_tarikh(tarikh),
            topik=topik,
            jam=", ".join(jam_parts) if jam_parts else "",
            hpk="HPK",
            catatan=catatan,
        ))

    return weeks


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN XLSX PARSER (combines Layer 1 + Layer 2)
# ═══════════════════════════════════════════════════════════════════════════

def parse_input_xlsx(file_bytes: bytes) -> tuple[UploadMetadata, list[WeekData]]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    meta_dict = _extract_metadata_anchored(ws, max_row=20)
    pattern_meta = _extract_metadata_patterns(ws, max_row=20)
    if pattern_meta:
        for key, val in pattern_meta.items():
            if val:
                meta_dict[key] = val

    metadata = UploadMetadata(
        program=str(meta_dict.get("program", "") or ""),
        semester=str(meta_dict.get("semester", "") or ""),
        tahun=str(meta_dict.get("tahun", "") or ""),
        pensyarah=str(meta_dict.get("pensyarah", "") or ""),
        ambilan=str(meta_dict.get("ambilan", "") or ""),
        jabatan=str(meta_dict.get("jabatan", "") or ""),
        kumpulan_diajar=meta_dict.get("kumpulan_diajar", []) if isinstance(meta_dict.get("kumpulan_diajar"), list) else _split_kumpulan(str(meta_dict.get("kumpulan_diajar", ""))),
        nama_kursus=str(meta_dict.get("nama_kursus", "") or ""),
        kod_kursus=str(meta_dict.get("kod_kursus", "") or ""),
        jumlah_kredit=str(meta_dict.get("jumlah_kredit", "") or ""),
    )

    weeks: list[WeekData] = []
    header_row = _find_header_row(ws)
    if header_row:
        col_map = _map_columns(ws, header_row)
        jam_col_map = _discover_jam_columns(ws, header_row)
        weeks = _extract_weeks_from_xlsx(ws, header_row, col_map, jam_col_map)

    if not weeks:
        weeks = [WeekData(minggu=i, hpk="HPK") for i in range(1, 15)]

    weeks.sort(key=lambda w: w.minggu)
    wb.close()

    print(f"[CLP XLSX Parse] nama_kursus='{metadata.nama_kursus}', kod='{metadata.kod_kursus}', "
          f"pensyarah='{metadata.pensyarah}', kumpulan={metadata.kumpulan_diajar}, weeks={len(weeks)}")

    return metadata, weeks


# ═══════════════════════════════════════════════════════════════════════════
#  PDF PARSER
# ═══════════════════════════════════════════════════════════════════════════

def parse_input_pdf(file_bytes: bytes) -> tuple[UploadMetadata, list[WeekData]]:
    metadata = UploadMetadata()
    weeks: list[WeekData] = []

    try:
        pdf = pdfplumber.open(BytesIO(file_bytes))
    except Exception:
        return metadata, [WeekData(minggu=i, hpk="HPK") for i in range(1, 15)]

    full_text = ""
    for page in pdf.pages:
        page_text = page.extract_text() or ""
        full_text += page_text + "\n"
    pdf.close()

    if len(full_text.strip()) < 50:
        return metadata, [WeekData(minggu=i, hpk="HPK") for i in range(1, 15)]

    lines = [ln.strip() for ln in full_text.split("\n") if ln.strip()]

    def _find_value_after_anchor(anchor: str) -> str:
        for line in lines:
            idx = line.lower().find(anchor.lower())
            if idx == -1:
                continue
            after = line[idx + len(anchor):]
            after = after.lstrip(" \t:")
            if after:
                return _clean(after)
        return ""

    for anchors, field_name in _METADATA_ANCHORS:
        for anchor in anchors:
            value = _find_value_after_anchor(anchor)
            if value:
                for stop_anchors, _ in _METADATA_ANCHORS:
                    for stop in stop_anchors:
                        stop_idx = value.lower().find(stop)
                        if stop_idx > 0:
                            value = value[:stop_idx].strip().rstrip(":")
                            break
                if not value:
                    continue
                if field_name == "kumpulan_diajar":
                    if not metadata.kumpulan_diajar:
                        metadata.kumpulan_diajar = _split_kumpulan(value)
                elif not getattr(metadata, field_name, None):
                    setattr(metadata, field_name, value)
                break

    if not metadata.kod_kursus:
        m = _KOD_KURSUS_PATTERN.search(full_text[:2000])
        if m:
            metadata.kod_kursus = m.group(1)

    if not metadata.pensyarah:
        m = _NAMA_PATTERN.search(full_text[:2000])
        if m:
            metadata.pensyarah = _clean(m.group(1))

    seen_weeks: set[int] = set()
    all_tables = []
    try:
        pdf2 = pdfplumber.open(BytesIO(file_bytes))
        for page in pdf2.pages[:5]:
            tables = page.extract_tables()
            if tables:
                all_tables.extend(tables)
        pdf2.close()
    except Exception:
        pass

    for table in all_tables:
        if not table or len(table) < 2:
            continue
        for row in table:
            if not row or len(row) < 3:
                continue
            try:
                week_num = int(str(row[0]).strip())
            except (ValueError, TypeError):
                continue
            if week_num < 1 or week_num > 20 or week_num in seen_weeks:
                continue
            topik = ""
            tarikh = ""
            for cell in row[1:]:
                cell_str = _clean(cell)
                if not cell_str:
                    continue
                if re.search(r'\d{1,2}\s*[-\u2013]\s*\d{1,2}', cell_str) and not tarikh:
                    tarikh = cell_str
                elif len(cell_str) > len(topik):
                    topik = cell_str

            topik_lower = topik.lower()
            if any(kw in topik_lower for kw in ["cuti", "peperiksaan", "ulangkaji", "pertengahan semester"]):
                if "akhir semester" in topik_lower:
                    topik = "CUTI AKHIR SEMESTER IPG"
                elif "pertengahan" in topik_lower or "cuti" in topik_lower:
                    topik = "CUTI PERTENGAHAN SEMESTER IPG"
                elif "ulangkaji" in topik_lower:
                    topik = "MINGGU ULANGKAJI"
                elif "peperiksaan" in topik_lower:
                    topik = "PEPERIKSAAN AKHIR"

            if re.match(r'^\d{1,2}\s+\w+\s*[-\u2013]\s*\d{1,2}\s+\w+\s+\d{4}$', topik.strip()):
                continue

            seen_weeks.add(week_num)
            weeks.append(WeekData(
                minggu=week_num,
                tarikh=_format_tarikh(tarikh),
                topik=topik,
                hpk="HPK",
            ))

    if not weeks:
        weeks = [WeekData(minggu=i, hpk="HPK") for i in range(1, 15)]

    weeks.sort(key=lambda w: w.minggu)
    return metadata, weeks


# ---------------------------------------------------------------------------
#  Tarikh gap-fill
# ---------------------------------------------------------------------------

_MALAY_MONTHS = {
    "januari": 1, "februari": 2, "mac": 3, "april": 4, "mei": 5, "jun": 6,
    "julai": 7, "ogos": 8, "september": 9, "oktober": 10, "november": 11, "disember": 12,
}


def _parse_tarikh_to_date(tarikh: str):
    parts = tarikh.strip().split()
    if len(parts) < 3:
        return None
    try:
        day = int(parts[0])
        month = _MALAY_MONTHS.get(parts[1].lower())
        year = int(parts[2])
        if month:
            return dt_module.date(year, month, day)
    except (ValueError, IndexError):
        pass
    return None


def _date_to_tarikh(d) -> str:
    month_names = {
        1: "Januari", 2: "Februari", 3: "Mac", 4: "April", 5: "Mei", 6: "Jun",
        7: "Julai", 8: "Ogos", 9: "September", 10: "Oktober", 11: "November", 12: "Disember",
    }
    return f"{d.day} {month_names[d.month]} {d.year}"


def _fill_tarikh_gaps(weeks: list[WeekData]) -> None:
    if not weeks:
        return
    known: dict[int, object] = {}
    for w in weeks:
        if w.tarikh:
            d = _parse_tarikh_to_date(w.tarikh)
            if d:
                known[w.minggu] = d
    if not known:
        return
    for w in weeks:
        if w.tarikh:
            continue
        closest_week = min(known.keys(), key=lambda k: abs(k - w.minggu))
        closest_date = known[closest_week]
        delta_weeks = w.minggu - closest_week
        estimated = closest_date + dt_module.timedelta(weeks=delta_weeks)
        w.tarikh = _date_to_tarikh(estimated)


# ═══════════════════════════════════════════════════════════════════════════
#  LAYER 3: LLM-ASSISTED HYBRID PARSER
# ═══════════════════════════════════════════════════════════════════════════

async def parse_with_ai(file_bytes: bytes, filename: str) -> tuple[UploadMetadata, list[WeekData]]:
    filename_lower = filename.lower()

    if filename_lower.endswith(".pdf"):
        raw_text = _extract_text_from_pdf(file_bytes)
    else:
        raw_text = _extract_text_from_xlsx(file_bytes)

    if not raw_text or len(raw_text.strip()) < 50:
        if filename_lower.endswith(".pdf"):
            return parse_input_pdf(file_bytes)
        else:
            return parse_input_xlsx(file_bytes)

    data = await extract_file_content(raw_text)
    meta_raw = data.get("metadata", {}) or {}

    meta_dict: dict[str, str] = {}
    for k, v in meta_raw.items():
        if isinstance(k, str):
            meta_dict[k.strip().lower().replace(" ", "_")] = v

    def _ai_get(keys: list[str], substrings: list[str] | None = None, default=""):
        for key in keys:
            kn = key.strip().lower().replace(" ", "_")
            if kn in meta_dict and meta_dict[kn] not in (None, ""):
                return meta_dict[kn]
        if substrings:
            for k, v in meta_dict.items():
                if all(s in k.lower() for s in substrings) and v not in (None, ""):
                    return v
        return default

    raw_kumpulan = _ai_get(["kumpulan_diajar", "kumpulan", "kumpulan_ajar"], ["kumpulan"], default=[])
    if isinstance(raw_kumpulan, str):
        kumpulan_list = _split_kumpulan(raw_kumpulan)
    elif isinstance(raw_kumpulan, list):
        kumpulan_list = [str(g).strip() for g in raw_kumpulan if str(g).strip()]
    else:
        kumpulan_list = []

    metadata = UploadMetadata(
        program=str(_ai_get(["program"], ["program"]) or ""),
        semester=str(_ai_get(["semester"], ["semester"]) or ""),
        tahun=str(_ai_get(["tahun", "tahun_akademik"], ["tahun"]) or ""),
        pensyarah=str(_ai_get(["pensyarah", "nama_pensyarah"], ["pensyarah"]) or ""),
        ambilan=str(_ai_get(["ambilan", "intake"], ["ambilan"]) or ""),
        jabatan=str(_ai_get(["jabatan", "unit"], ["jabatan"]) or ""),
        kumpulan_diajar=kumpulan_list,
        nama_kursus=str(_ai_get(["nama_kursus", "nama_mata_pelajaran"], ["kursus"]) or ""),
        kod_kursus=str(_ai_get(["kod_kursus", "kod"], ["kod"]) or ""),
        jumlah_kredit=str(_ai_get(["jumlah_kredit", "kredit"], ["kredit"]) or ""),
    )

    weeks: list[WeekData] = []
    for w in data.get("weeks", []):
        try:
            minggu = int(w.get("minggu", 0))
            if minggu < 1 or minggu > 19:
                continue
            raw_topik = str(w.get("topik", ""))
            raw_lower = raw_topik.strip().lower()
            if any(kw in raw_lower for kw in ["cuti", "peperiksaan", "ulangkaji", "pertengahan semester"]):
                if "akhir semester" in raw_lower:
                    raw_topik = "CUTI AKHIR SEMESTER IPG"
                elif "pertengahan" in raw_lower or ("cuti" in raw_lower and "akhir" not in raw_lower):
                    raw_topik = "CUTI PERTENGAHAN SEMESTER IPG"
                elif "ulangkaji" in raw_lower:
                    raw_topik = "MINGGU ULANGKAJI"
                elif "peperiksaan" in raw_lower:
                    raw_topik = "PEPERIKSAAN AKHIR"
            weeks.append(WeekData(
                minggu=minggu,
                tarikh=_format_tarikh(str(w.get("tarikh", ""))),
                topik=raw_topik,
                jam="",
                hpk="HPK",
                catatan=str(w.get("catatan", "")),
            ))
        except (ValueError, TypeError):
            continue

    if not weeks:
        raise ValueError("AI tidak berjaya mengekstrak data minggu")

    weeks.sort(key=lambda w: w.minggu)

    # Deterministic overlay
    if not filename_lower.endswith(".pdf"):
        try:
            det_meta, det_weeks = parse_input_xlsx(file_bytes)
            jam_map = {rw.minggu: rw.jam for rw in det_weeks if rw.jam}
            for week in weeks:
                if week.minggu in jam_map:
                    week.jam = jam_map[week.minggu]
            if det_meta.nama_kursus:
                metadata.nama_kursus = det_meta.nama_kursus
            if det_meta.kod_kursus:
                metadata.kod_kursus = det_meta.kod_kursus
            if det_meta.pensyarah:
                metadata.pensyarah = det_meta.pensyarah
            if det_meta.kumpulan_diajar:
                metadata.kumpulan_diajar = det_meta.kumpulan_diajar
            for field in ["program", "semester", "tahun", "jabatan", "ambilan", "jumlah_kredit"]:
                if not getattr(metadata, field) and getattr(det_meta, field, None):
                    setattr(metadata, field, getattr(det_meta, field))
            tarikh_map = {rw.minggu: rw.tarikh for rw in det_weeks if rw.tarikh}
            for week in weeks:
                if not week.tarikh and week.minggu in tarikh_map:
                    week.tarikh = tarikh_map[week.minggu]
        except Exception as e:
            print(f"[CLP Hybrid] Deterministic XLSX overlay failed: {e}")
    else:
        try:
            det_meta, _ = parse_input_pdf(file_bytes)
            if det_meta.nama_kursus:
                metadata.nama_kursus = det_meta.nama_kursus
            if det_meta.kod_kursus:
                metadata.kod_kursus = det_meta.kod_kursus
            if det_meta.pensyarah:
                metadata.pensyarah = det_meta.pensyarah
            if det_meta.kumpulan_diajar:
                metadata.kumpulan_diajar = det_meta.kumpulan_diajar
        except Exception as e:
            print(f"[CLP Hybrid] Deterministic PDF overlay failed: {e}")

    _fill_tarikh_gaps(weeks)
    return metadata, weeks


# ═══════════════════════════════════════════════════════════════════════════
#  TEXT EXTRACTION HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _extract_text_from_xlsx(file_bytes: bytes) -> str:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    lines: list[str] = []

    weekly_start = ws.max_row + 1
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=False):
        for cell in row:
            if cell.value and "minggu" in str(cell.value).strip().lower():
                weekly_start = cell.row
                break
        if weekly_start <= ws.max_row:
            break

    lines.append("=== MAKLUMAT KURSUS (METADATA) ===")
    for row in ws.iter_rows(min_row=1, max_row=min(weekly_start - 1, ws.max_row), values_only=False):
        cells = [str(c.value).strip() for c in row if c.value is not None and str(c.value).strip()]
        if cells:
            lines.append(" : ".join(cells))

    lines.append("\n=== DATA MINGGUAN ===")
    for row in ws.iter_rows(min_row=weekly_start, values_only=False):
        cells = [str(c.value).strip() for c in row if c.value is not None and str(c.value).strip()]
        if cells:
            lines.append(" | ".join(cells))

    wb.close()
    return "\n".join(lines)


def _extract_text_from_pdf(file_bytes: bytes) -> str:
    try:
        pdf = pdfplumber.open(BytesIO(file_bytes))
        parts = [page.extract_text() or "" for page in pdf.pages]
        pdf.close()
        return "\n".join(parts)
    except Exception:
        return ""
